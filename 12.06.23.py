#todo Регулярные выражения
#1 Напишите программу, в которой возвращаются домены из списка e-mail адресов.
'''
import re

emails = input("Введите список e-mail адресов через запятую: ").split(",")

domains = []

for email in emails:
    match = re.search(r'(?<=@)[^.]+\.[^,]+', email)
    if match:
        domains.append(match.group())

print("Домены: ", ", ".join(domains))
'''
#2 Напишите программу, в которой извлекаются слова, начинающиеся на гласную букву.
'''
import re

text = input("Введите текст: ")

matches = re.findall(r'\b[aeiouAEIOUуеыаоэяию]\w+', text)

print("Слова, начинающиеся на гласную букву: ", ", ".join(matches))
'''
#3 Напишите программу, в которой разбивается строка по нескольким разделителям
'''
import re

string = input("Введите строку: ")

delimiters = ",;:.- "
words = re.split('[' + delimiters + ']', string)

print("Слова из строки: ", words)
'''
#todo Модули и пакеты
#Спроектировать программу для определения победителя на выборах.
'''
def find_winner(candidates):
    votes = {}
    for candidate in candidates:
        if candidate in votes:
            votes[candidate] += 1
        else:
            votes[candidate] = 1
    max_votes = max(votes.values())
    winners = [k for k, v in votes.items() if v == max_votes]
    if len(winners) > 1:
        winners = sorted(winners, key=lambda x: (len(x), x))
    return winners[0], max_votes

candidates = ['Пешая', 'Аскаров', 'Бекмуханов', 'Ернур', 'Пешая', 'Карим', 'Шаримазданов', 'Ернур', 'Пешая',]
winner, votes = find_winner(candidates)
print(f"Победитель выборов: {winner} с количеством голосов: {votes}")
'''

#TODO задание 2

from openpyxl import Workbook, load_workbook
import re
def extra():
    col1 = []
    workbook = load_workbook('Лист1.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col1.append(sheet.cell(row=i, column=1).value)

    col2 = []
    workbook = load_workbook('Лист2.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col2.append(sheet.cell(row=i, column=2).value)

    col3 = []
    workbook = load_workbook('Лист3.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col3.append(sheet.cell(row=i, column=3).value)

    matrix = [col1, col2, col3]

    combined_file = Workbook()
    combined_sheet = combined_file.active
    row_i = 0
    for i in matrix:
        row_i += 1
        col_i = 0
        for j in i:
            col_i += 1
            combined_sheet.cell(row=row_i, column=col_i, value=j)
    combined_file.save('combined_file.xlsx')